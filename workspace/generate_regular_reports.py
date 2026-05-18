import asyncio
import os
import json
import math
import pandas as pd
import asyncpg
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

load_dotenv('.env')
# DB Configs
DB_CONFIGS = [
    {"name": "REGULAR", "host": os.getenv("REGULAR_LMS_DB_HOST"), "port": int(os.getenv("REGULAR_LMS_DB_PORT", "54321")), "database": os.getenv("REGULAR_LMS_DB_NAME"), "user": os.getenv("REGULAR_LMS_DB_USER"), "password": os.getenv("REGULAR_LMS_DB_PASSWORD")},
    {"name": "CGC", "host": os.getenv("REGULAR_CGC_LMS_DB_HOST"), "port": int(os.getenv("REGULAR_CGC_LMS_DB_PORT", "54321")), "database": os.getenv("REGULAR_CGC_LMS_DB_NAME"), "user": os.getenv("REGULAR_CGC_LMS_DB_USER"), "password": os.getenv("REGULAR_CGC_LMS_DB_PASSWORD")},
    {"name": "AMITY", "host": os.getenv("REGULAR_AMITY_LMS_DB_HOST"), "port": int(os.getenv("REGULAR_AMITY_LMS_DB_PORT", "54321")), "database": os.getenv("REGULAR_AMITY_LMS_DB_NAME"), "user": os.getenv("REGULAR_AMITY_LMS_DB_USER"), "password": os.getenv("REGULAR_AMITY_LMS_DB_PASSWORD")}
]

CONFIG_FILE = 'regular_report_config.json'
with open(CONFIG_FILE, 'r') as f:
    config = json.load(f)

now_utc = datetime.utcnow()
now_ist = now_utc + timedelta(hours=5, minutes=30)
report_date = now_ist - timedelta(days=1) if now_ist.hour < 6 else now_ist

FTD_DATE = report_date.strftime('%Y-%m-%d')
MTD_START = report_date.replace(day=1).strftime('%Y-%m-%d')
MTD_END = FTD_DATE
MONTH_LABEL = report_date.strftime('%b')
WEEK_START = config.get("target_period", {}).get("start_date")
WEEK_END = config.get("target_period", {}).get("end_date")

ADM_SQL = "SELECT DISTINCT ON (s.student_id, uc.course_id) s.student_id, uc.university_name AS college_name, csj.created_at AT TIME ZONE 'Asia/Kolkata' AS created_at FROM students s JOIN course_status_journeys csj ON s.student_id = csj.student_id JOIN university_courses uc ON csj.course_id = uc.course_id WHERE csj.course_status = 'Admission' AND COALESCE(csj.fee_type, '') NOT ILIKE '%partial%' AND s.student_id IN (SELECT student_id FROM students) ORDER BY s.student_id, uc.course_id, csj.created_at ASC;"
FORM_SQL = "SELECT DISTINCT ON (s.student_id, csj.course_id) s.student_id, uc.university_name AS college_name, csj.created_at AT TIME ZONE 'Asia/Kolkata' AS created_at FROM students s JOIN course_status_journeys csj ON s.student_id = csj.student_id JOIN university_courses uc ON csj.course_id = uc.course_id WHERE csj.course_status in ('Form Submitted – Portal Pending', 'Form Submitted – Completed', 'Walkin Completed', 'Exam/Interview Scheduled', 'Offer Letter/Results Pending', 'Offer Letter/Results Released', 'Ready For Admission') AND s.student_id IN (SELECT student_id FROM students) ORDER BY s.student_id, csj.course_id, csj.created_at Asc;"

async def get_data():
    all_adm, all_form = [], []
    for db in DB_CONFIGS:
        conn = await asyncpg.connect(host=db['host'], port=db['port'], database=db['database'], user=db['user'], password=db['password'])
        all_adm.append(pd.DataFrame([dict(r) for r in await conn.fetch(ADM_SQL)]))
        all_form.append(pd.DataFrame([dict(r) for r in await conn.fetch(FORM_SQL)]))
        await conn.close()
    
    def norm(name): return "Amity University (All Campuses)" if name and "Amity" in name else name
    df_adm = pd.concat(all_adm).assign(college_name=lambda x: x['college_name'].apply(norm))
    df_form = pd.concat(all_form).assign(college_name=lambda x: x['college_name'].apply(norm))
    return df_adm, df_form

def style_ws(ws):
    header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25
        for cell in col:
            cell.border = border
            if cell.row == 1:
                cell.fill, cell.font = header_fill, header_font
                cell.alignment = Alignment(horizontal='center')
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
        if row[0].value == "Total":
            for cell in row: cell.font = Font(bold=True); cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

def pct(a, t):
    if isinstance(a, float) and a.is_integer(): a = int(a)
    if isinstance(t, float) and t.is_integer(): t = int(t)
    return f"{(a/t*100):.1f}%" if t else "0.0%"

def generate():
    df_adm, df_form = asyncio.run(get_data())
    wb = Workbook()
    wb.remove(wb.active)
    COLLEGE_TARGETS = config.get("college_targets", {})

    for mode in ["Admissions Data", "Forms Data"]:
        ws = wb.create_sheet(mode)
        ws.append(["College", "YTD Ach", f"{MONTH_LABEL} Target", f"{MONTH_LABEL} Ach", f"{MONTH_LABEL} Ach %", "Week Target", "Week Ach", "Week Ach %", "FTD Target", "FTD Ach", "FTD Ach %"])
        df = df_adm if "Admissions" in mode else df_form
        df['date'] = df['created_at'].dt.strftime('%Y-%m-%d')
        
        totals = [0]*7 # YTD, AprT, AprA, W5T, W5A, FTDT, FTDA
        for col, tgts in COLLEGE_TARGETS.items():
            key_prefix = "admission" if "Admissions" in mode else "forms"
            apr_t = tgts.get(f"{key_prefix}_apr", 0)
            w5_t = tgts.get(f"{key_prefix}_week", 0)
            # Daily target must be rounded UP, not Python's round().
            # Example: weekly target 29 / 7 = 4.14 should show FTD Target 5, never 4.
            ftd_t = math.ceil(w5_t / 7) if w5_t else 0
            
            ytd_a = len(df[df['college_name'] == col])
            apr_a = len(df[(df['college_name'] == col) & (df['date'].between(MTD_START, MTD_END))])
            w5_a = len(df[(df['college_name'] == col) & (df['date'].between(WEEK_START, WEEK_END))])
            ftd_a = len(df[(df['college_name'] == col) & (df['date'] == FTD_DATE)])
            
            ws.append([col, ytd_a, apr_t, apr_a, pct(apr_a, apr_t), w5_t, w5_a, pct(w5_a, w5_t), ftd_t, ftd_a, pct(ftd_a, ftd_t)])
            for i, v in enumerate([ytd_a, apr_t, apr_a, w5_t, w5_a, ftd_t, ftd_a]): totals[i] += v
            
        ws.append(["Total", totals[0], totals[1], totals[2], pct(totals[2], totals[1]), totals[3], totals[4], pct(totals[4], totals[3]), totals[5], totals[6], pct(totals[6], totals[5])])
        style_ws(ws)
    
    wb.save('Daily_Regular_LMS_Reports.xlsx')

if __name__ == "__main__": generate()
