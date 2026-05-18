import asyncio
import os
import pandas as pd
import asyncpg
import base64
import requests
from datetime import datetime
import pytz
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Load environment variables (absolute path for cron)
load_dotenv('/home/hermeswebui/workspace/.env')

DB_HOST = os.getenv("ONLINE_LMS_DB_HOST")
DB_PORT = int(os.getenv("ONLINE_LMS_DB_PORT", "54321"))
DB_NAME = os.getenv("ONLINE_LMS_DB_NAME")
DB_USER = os.getenv("ONLINE_LMS_DB_USER")
DB_PASSWORD = os.getenv("ONLINE_LMS_DB_PASSWORD")
WHAPI_TOKEN = os.getenv("WHAPI_TOKEN")
WHATSAPP_GROUP = os.getenv("WHATSAPP_GROUP")

# Get today's date bounds in IST (00:00:00 to 23:59:59)
ist = pytz.timezone('Asia/Kolkata')
today_ist = datetime.now(ist)
start_date = today_ist.strftime('%Y-%m-%d 00:00:00')
end_date = today_ist.strftime('%Y-%m-%d 23:59:59')
display_date = today_ist.strftime('%d %b %Y')

QUERY = f"""
SELECT DISTINCT ON (s.student_id, uc.course_id)
    s.student_id,
    csj.deposit_amount AS fee_deposit,
    c.counsellor_name
FROM students s
JOIN course_status_journeys csj ON s.student_id = csj.student_id
JOIN university_courses uc ON csj.course_id = uc.course_id
LEFT JOIN counsellors c ON s.assigned_counsellor_id = c.counsellor_id
WHERE csj.course_status = 'Admission'
  AND csj.fee_type NOT IN ('partial paid', 'Partially Paid', 'Partial Done','Partial Paid')
  AND s.student_id IN (SELECT student_id FROM students)
  AND (csj.created_at AT TIME ZONE 'Asia/Kolkata') >= '{start_date}'
  AND (csj.created_at AT TIME ZONE 'Asia/Kolkata') <= '{end_date}'
"""

SUPERVISORS = {
    'Varun': ['Abhishek Kamat', 'Abhishek Swami', 'Amandeep', 'Amit Kumar', 'Anishika Prashar', 'Ankita Singh Chauhan', 'Arnav Upadhyay', 'Avneet'],
    'Sunil': ['Awantika Singh', 'Chandni Kumari', 'Divya Goyal', 'Himanshi', 'Himanshi Baliyan', 'Kumud Kumar', 'Laxminaryan', 'Mayank Jain'],
    'Vishal': ['Mousumi Khatun', 'Om  Sharma', 'Prashant', 'Prateek Chauhan', 'Preeti Lohiya', 'Purushottam Kumar', 'Raja Ram', 'Rupal Tiwari'],
    'Siddhartha': ['Siddarth Kumar', 'Suhani Raj Gupta', 'Sumit Yadav', 'Tanisha Kulshrestha', 'Tanya Singh', 'Tanya sharma', 'Vijay Kumar', 'Vikas', 'Vishwajeet']
}

def get_supervisor(counselor_name):
    for sup, counselors in SUPERVISORS.items():
        if counselor_name in counselors:
            return sup
    return 'Other'

async def main():
    conn = await asyncpg.connect(
        user=DB_USER, password=DB_PASSWORD, database=DB_NAME,
        host=DB_HOST, port=DB_PORT
    )
    records = await conn.fetch(QUERY)
    await conn.close()
    
    headers_req = {
        "accept": "application/json",
        "authorization": f"Bearer {WHAPI_TOKEN}",
        "content-type": "application/json"
    }
    
    df = pd.DataFrame([dict(r) for r in records])
    if df.empty:
        print("No admissions recorded today.")
        payload = {
            "to": WHATSAPP_GROUP,
            "body": f"📊 *Daily Admission Report | {display_date}*\n\nThere were no new admissions recorded today."
        }
        requests.post("https://gate.whapi.cloud/messages/text", headers=headers_req, json=payload)
        return

    df['is_ftd'] = pd.to_numeric(df['fee_deposit'], errors='coerce').fillna(0) > 0
    df['counsellor_name'] = df['counsellor_name'].fillna('Unassigned')
    
    counselor_stats = df.groupby('counsellor_name').agg(
        achieve=('student_id', 'count'),
        ftd=('is_ftd', 'sum')
    ).reset_index()
    
    counselor_stats['supervisor'] = counselor_stats['counsellor_name'].apply(get_supervisor)
    
    excel_rows = [["Supervisor", "Counselor", "Achieve", "FTD Achievement"]]
    grand_achieve = 0
    grand_ftd = 0
    
    for sup in ['Varun', 'Sunil', 'Vishal', 'Siddhartha', 'Other']:
        sup_data = counselor_stats[counselor_stats['supervisor'] == sup]
        if sup_data.empty:
            continue
            
        sup_achieve = 0
        sup_ftd = 0
        sup_data = sup_data.sort_values('counsellor_name')
        
        for _, row in sup_data.iterrows():
            ach = row['achieve']
            ftd = row['ftd']
            excel_rows.append([sup, row['counsellor_name'], ach, ftd])
            sup_achieve += ach
            sup_ftd += ftd
            grand_achieve += ach
            grand_ftd += ftd
            
        excel_rows.append([f"{sup} Total", "", sup_achieve, sup_ftd])
        
    excel_rows.append(["Grand Total", "", grand_achieve, grand_ftd])
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Admissions {display_date}"
    
    for r in excel_rows:
        ws.append(r)
        
    # Styling
    header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
    grand_total_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        is_header = (row_idx == 1)
        val = str(row[0].value) if row[0].value else ""
        is_grand_total = (val == "Grand Total")
        is_sup_total = (val.endswith(" Total") and not is_grand_total)
        
        for col_idx, cell in enumerate(row):
            cell.border = thin_border
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
            elif is_grand_total:
                cell.fill = grand_total_fill
                cell.font = bold_font
            elif is_sup_total:
                cell.fill = total_fill
                cell.font = bold_font
                
            if col_idx >= 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    excel_path = f"/home/hermeswebui/workspace/Daily_Admission_Report.xlsx"
    wb.save(excel_path)
    
    # Send via WHAPI
    with open(excel_path, "rb") as f:
        b64_content = base64.b64encode(f.read()).decode('utf-8')
        
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    media_payload = f"data:{mime_type};name=Daily_Admission_Report_{display_date.replace(' ', '_')}.xlsx;base64,{b64_content}"
    
    payload = {
        "to": WHATSAPP_GROUP,
        "media": media_payload,
        "caption": f"📊 *Daily Admission Report | {display_date}*\n\nHere is the end-of-day target vs achievement summary."
    }
    
    resp = requests.post("https://gate.whapi.cloud/messages/document", headers=headers_req, json=payload)
    print("Report dispatched successfully.")

if __name__ == "__main__":
    asyncio.run(main())