import asyncio
import os
import sys
import json
import pandas as pd
import asyncpg
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE_DIR = os.environ.get('WORKSPACE_DIR', os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, '.env'))

DB_HOST = os.getenv("ONLINE_LMS_DB_HOST")
DB_PORT = int(os.getenv("ONLINE_LMS_DB_PORT", "54321"))
DB_NAME = os.getenv("ONLINE_LMS_DB_NAME")
DB_USER = os.getenv("ONLINE_LMS_DB_USER")
DB_PASSWORD = os.getenv("ONLINE_LMS_DB_PASSWORD")

# Load configuration and targets
CONFIG_FILE = os.path.join(BASE_DIR, 'report_config.json')
with open(CONFIG_FILE, 'r') as f:
    config = json.load(f)

# -------------------------------------------------------------
# DYNAMIC DATE LOGIC FOR FUTURE-PROOFING
# -------------------------------------------------------------
# If current time is before 6 AM IST, we consider "today's report" to be for yesterday.
now_utc = datetime.utcnow()
now_ist = now_utc + timedelta(hours=5, minutes=30)

if now_ist.hour < 6:
    report_date = now_ist - timedelta(days=1)
else:
    report_date = now_ist

FTD_DATE = report_date.strftime('%Y-%m-%d')

# 2. MTD is ALWAYS the 1st of the current month to the report date
MTD_START = report_date.replace(day=1).strftime('%Y-%m-%d')
MTD_END = FTD_DATE

# 3. Weekly/Target Period is pulled from the config file. 
WEEK_START = config.get("target_period", {}).get("start_date", (report_date - timedelta(days=report_date.weekday())).strftime('%Y-%m-%d'))
WEEK_END = config.get("target_period", {}).get("end_date", FTD_DATE)

print(f"Generating report for FTD: {FTD_DATE}, WEEK: {WEEK_START} to {WEEK_END}, MTD: {MTD_START} to {MTD_END}")

# Load Targets
SUPERVISOR_TARGETS = config.get("supervisor_targets", {})
REVENUE_TARGETS = config.get("counsellor_targets", {})

async def get_data():
    conn = await asyncpg.connect(
        host=DB_HOST, port=DB_PORT, database=DB_NAME, user=DB_USER, password=DB_PASSWORD
    )
    
    # Active counsellors
    couns_data = []
    for sup, couns_dict in REVENUE_TARGETS.items():
        for couns_name in couns_dict.keys():
            couns_data.append({'supervisor_name': sup, 'counsellor_name': couns_name})
    df_couns = pd.DataFrame(couns_data)
    
    adm_query = """
    SELECT 
        s.student_id,
        uc.university_name AS college_name,
        COALESCE(csj.deposit_amount, 0) AS fee_deposit,
        -- Use journey counsellor. If they have no supervisor above them (jsup is NULL),
        -- they are a supervisor themselves; fall back to current assigned counsellor.
        CASE WHEN jsup.counsellor_name IS NULL THEN c_fallback.counsellor_name ELSE jc.counsellor_name END AS counsellor_name,
        -- Supervisor: use journey counsellor's supervisor; if journey counsellor IS a supervisor, use their name.
        COALESCE(jsup.counsellor_name, jc.counsellor_name) AS supervisor_name,
        csj.created_at AT TIME ZONE 'Asia/Kolkata' AS created_at
    FROM students s
    JOIN course_status_journeys csj ON s.student_id = csj.student_id
    JOIN university_courses uc ON csj.course_id = uc.course_id
    LEFT JOIN counsellors jc ON csj.counsellor_id = jc.counsellor_id
    LEFT JOIN counsellors jsup ON jc.assigned_to = jsup.counsellor_id
    LEFT JOIN counsellors c_fallback ON s.assigned_counsellor_id = c_fallback.counsellor_id
    WHERE csj.course_status = 'Admission'
      AND COALESCE(csj.fee_type, '') NOT ILIKE '%partial%'
    """
    
    form_query = """
    SELECT 
        s.student_id,
        uc.university_name AS college_name,
        csj.created_at AT TIME ZONE 'Asia/Kolkata' AS created_at
    FROM students s
    JOIN course_status_journeys csj ON s.student_id = csj.student_id
    JOIN university_courses uc ON csj.course_id = uc.course_id
    WHERE csj.course_status = 'Application'
    """
    
    adm_rows = await conn.fetch(adm_query)
    form_rows = await conn.fetch(form_query)
    await conn.close()
    
    df_adm = pd.DataFrame([dict(r) for r in adm_rows])
    df_form = pd.DataFrame([dict(r) for r in form_rows])

    # Normalize names before merging DB rows with report_config.json.
    # The LMS has hidden trailing spaces in some counsellor names (e.g. "Vishwajeet ").
    # Without this, admissions/revenue exist in raw DB and college totals, but disappear
    # from counsellor/supervisor tables because pandas merge is exact-string based.
    for df in (df_couns, df_adm):
        if not df.empty:
            for col in ('supervisor_name', 'counsellor_name'):
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
    
    # Deduplicate to prevent double-counting
    if not df_adm.empty:
        df_adm = df_adm.sort_values(['student_id', 'college_name', 'created_at']).drop_duplicates(subset=['student_id', 'college_name'], keep='first')
    if not df_form.empty:
        df_form = df_form.sort_values(['student_id', 'college_name', 'created_at']).drop_duplicates(subset=['student_id', 'college_name'], keep='first')
    
    return df_couns, df_adm, df_form

def style_worksheet(ws):
    header_fill = PatternFill(start_color="004040", end_color="004040", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    subtotal_fill = PatternFill(start_color="E0F0F0", end_color="E0F0F0", fill_type="solid")
    grand_total_fill = PatternFill(start_color="004040", end_color="004040", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4

    for row in ws.iter_rows(min_row=2):
        is_subtotal = any(cell.value and 'Total' in str(cell.value) for cell in row)
        is_grand_total = any(cell.value == 'Grand Total' for cell in row)
        
        if is_subtotal and not is_grand_total:
            for cell in row:
                cell.fill = subtotal_fill
                cell.font = Font(bold=True)
        elif is_grand_total:
            for cell in row:
                cell.fill = grand_total_fill
                cell.font = Font(color="FFFFFF", bold=True)

def pct(achieved, target):
    achieved = 0 if pd.isna(achieved) else achieved
    target = 0 if pd.isna(target) else target
    if isinstance(achieved, float) and achieved.is_integer():
        achieved = int(achieved)
    if isinstance(target, float) and target.is_integer():
        target = int(target)
    if target == 0 or pd.isna(target): return "0.0%"
    return f"{(achieved / target * 100):.1f}%"

def generate_reports():
    df_couns, df_adm, df_form = asyncio.run(get_data())
    
    if not df_adm.empty:
        df_adm['date_str'] = df_adm['created_at'].dt.strftime('%Y-%m-%d')
    else:
        df_adm['date_str'] = pd.Series(dtype='object')

    if not df_form.empty:
        df_form['date_str'] = df_form['created_at'].dt.strftime('%Y-%m-%d')
    else:
        df_form['date_str'] = pd.Series(dtype='object')
    
    # APPLY DYNAMIC TIME PERIODS
    adm_ftd = df_adm[df_adm['date_str'] == FTD_DATE]
    form_ftd = df_form[df_form['date_str'] == FTD_DATE]
    
    adm_week = df_adm[df_adm['date_str'].between(WEEK_START, WEEK_END)]
    
    adm_mtd = df_adm[df_adm['date_str'].between(MTD_START, MTD_END)]
    form_mtd = df_form[df_form['date_str'].between(MTD_START, MTD_END)]

    # Guardrail: expose admissions that will not land in counsellor/supervisor tables
    # because the DB counsellor/supervisor pair is absent from report_config.json.
    # This catches future hidden-space or spelling mismatches before totals silently drift.
    known_pairs = set(map(tuple, df_couns[['supervisor_name', 'counsellor_name']].drop_duplicates().values.tolist())) if not df_couns.empty else set()
    unmatched_adm_week = pd.DataFrame()
    if not adm_week.empty:
        unmatched_adm_week = adm_week[~adm_week.apply(lambda r: (r.get('supervisor_name'), r.get('counsellor_name')) in known_pairs, axis=1)].copy()
        if not unmatched_adm_week.empty:
            print("WARNING: Admissions found for counsellor/supervisor pairs missing from report_config.json")
            print(unmatched_adm_week[['student_id', 'college_name', 'supervisor_name', 'counsellor_name', 'fee_deposit', 'created_at']].to_string(index=False))

    wb = Workbook()
    wb.remove(wb.active)

    sup_order = ['Varun', 'Sunil', 'Vishal Gaur', 'Siddarth Kumar']
    display_names = {'Varun': 'Varun', 'Sunil': 'Sunil', 'Vishal Gaur': 'Vishal', 'Siddarth Kumar': 'Siddhartha'}

    # 1. COUNSELLOR FEE COLLECTED (Target Period vs FTD)
    couns_rev_week = adm_week.groupby(['supervisor_name', 'counsellor_name'])['fee_deposit'].sum().reset_index(name='Achieved')
    couns_rev_ftd = adm_ftd.groupby(['supervisor_name', 'counsellor_name'])['fee_deposit'].sum().reset_index(name='FTD')
    
    df_c_rev = df_couns.merge(couns_rev_week, on=['supervisor_name', 'counsellor_name'], how='left')\
                       .merge(couns_rev_ftd, on=['supervisor_name', 'counsellor_name'], how='left').fillna(0)
    
    df_c_rev['Target'] = df_c_rev.apply(lambda r: REVENUE_TARGETS.get(r['supervisor_name'], {}).get(r['counsellor_name'], 0), axis=1)
    df_c_rev['Ach %'] = df_c_rev.apply(lambda r: pct(r['Achieved'], r['Target']), axis=1)

    ws1 = wb.create_sheet("Counsellor_Fee_Collected")
    ws1.append(['Supervisor', 'Counsellor', 'Target', 'Fee Collected', 'Ach %', 'FTD'])
    
    for sup in sup_order:
        s_data = df_c_rev[df_c_rev['supervisor_name'] == sup]
        if s_data.empty: continue
        for _, row in s_data.iterrows():
            ws1.append([display_names.get(sup, sup), row['counsellor_name'], row['Target'], row['Achieved'], row['Ach %'], row['FTD']])
        ws1.append([f'Total ({display_names.get(sup, sup)})', '', s_data['Target'].sum(), s_data['Achieved'].sum(), pct(s_data['Achieved'].sum(), s_data['Target'].sum()), s_data['FTD'].sum()])
    ws1.append(['Grand Total', '', df_c_rev['Target'].sum(), df_c_rev['Achieved'].sum(), pct(df_c_rev['Achieved'].sum(), df_c_rev['Target'].sum()), df_c_rev['FTD'].sum()])
    style_worksheet(ws1)

    # 2. SUPERVISOR FEE COLLECTED
    ws2 = wb.create_sheet("Supervisor_Fee_Collected")
    ws2.append(['Supervisor', 'Target', 'Fee Collected', 'Ach %', 'FTD'])
    for sup in sup_order:
        s_data = df_c_rev[df_c_rev['supervisor_name'] == sup]
        sup_tgt = SUPERVISOR_TARGETS.get(sup, 0)
        sup_ach = s_data['Achieved'].sum()
        sup_ftd = s_data['FTD'].sum()
        ws2.append([display_names.get(sup, sup), sup_tgt, sup_ach, pct(sup_ach, sup_tgt), sup_ftd])
    gt_tgt = sum(SUPERVISOR_TARGETS.values())
    gt_ach = df_c_rev['Achieved'].sum()
    gt_ftd = df_c_rev['FTD'].sum()
    ws2.append(['Grand Total', gt_tgt, gt_ach, pct(gt_ach, gt_tgt), gt_ftd])
    style_worksheet(ws2)

    # 3. COUNSELLOR ADMISSION
    couns_adm_week = adm_week.groupby(['supervisor_name', 'counsellor_name']).size().reset_index(name='Achieve')
    couns_adm_ftd = adm_ftd.groupby(['supervisor_name', 'counsellor_name']).size().reset_index(name='FTD')
    
    df_c_adm = df_couns.merge(couns_adm_week, on=['supervisor_name', 'counsellor_name'], how='left')\
                       .merge(couns_adm_ftd, on=['supervisor_name', 'counsellor_name'], how='left').fillna(0)

    ws3 = wb.create_sheet("Counsellor_Admission")
    ws3.append(['Supervisor', 'Counsellor', 'Achieve', 'FTD'])
    for sup in sup_order:
        s_data = df_c_adm[df_c_adm['supervisor_name'] == sup]
        if s_data.empty: continue
        for _, row in s_data.iterrows():
            ws3.append([display_names.get(sup, sup), row['counsellor_name'], row['Achieve'], row['FTD']])
        ws3.append([f'Total ({display_names.get(sup, sup)})', '', s_data['Achieve'].sum(), s_data['FTD'].sum()])
    ws3.append(['Grand Total', '', df_c_adm['Achieve'].sum(), df_c_adm['FTD'].sum()])
    style_worksheet(ws3)

    # 4. SUPERVISOR ADMISSION
    ws4 = wb.create_sheet("Supervisor_Admission")
    ws4.append(['Supervisor', 'Achieve', 'FTD'])
    for sup in sup_order:
        s_data = df_c_adm[df_c_adm['supervisor_name'] == sup]
        ws4.append([display_names.get(sup, sup), s_data['Achieve'].sum(), s_data['FTD'].sum()])
    ws4.append(['Grand Total', df_c_adm['Achieve'].sum(), df_c_adm['FTD'].sum()])
    style_worksheet(ws4)

    # 5. COLLEGE PERFORMANCE
    if not df_adm.empty:
        df_adm['col_norm'] = df_adm['college_name'].str.lower().str.replace(' online', ' University Online').str.replace(' university university', ' University')
    
    uni_ytd_f = df_form.groupby('college_name').size().reset_index(name='YTD Forms')
    uni_ytd_a = df_adm.groupby('college_name').size().reset_index(name='YTD Admissions')
    
    uni_mtd_f = form_mtd.groupby('college_name').size().reset_index(name='MTD Forms')
    uni_mtd_a = adm_mtd.groupby('college_name').size().reset_index(name='MTD Admissions')
    
    uni_ftd_f = form_ftd.groupby('college_name').size().reset_index(name='FTD Forms')
    uni_ftd_a = adm_ftd.groupby('college_name').size().reset_index(name='FTD Admissions')

    df_uni = uni_ytd_f.merge(uni_ytd_a, on='college_name', how='outer')\
                      .merge(uni_mtd_f, on='college_name', how='outer')\
                      .merge(uni_mtd_a, on='college_name', how='outer')\
                      .merge(uni_ftd_f, on='college_name', how='outer')\
                      .merge(uni_ftd_a, on='college_name', how='outer').fillna(0)

    df_uni = df_uni.sort_values('YTD Forms', ascending=False)
    
    ws5 = wb.create_sheet("College_Performance")
    ws5.append(['Colleges', 'YTD Forms', 'YTD Admissions', 'YTD F2A %', 'MTD Forms', 'MTD Admissions', 'MTD F2A %', 'FTD Forms', 'FTD Admissions', 'FTD F2A %'])
    
    for _, row in df_uni.iterrows():
        c_name = row['college_name'].title().replace('Online', 'Online').replace('Noida', 'Noida')
        ws5.append([
            c_name, 
            int(row['YTD Forms']), int(row['YTD Admissions']), pct(row['YTD Admissions'], row['YTD Forms']),
            int(row['MTD Forms']), int(row['MTD Admissions']), pct(row['MTD Admissions'], row['MTD Forms']),
            int(row['FTD Forms']), int(row['FTD Admissions']), pct(row['FTD Admissions'], row['FTD Forms'])
        ])
    
    ws5.append([
        'Total', 
        int(df_uni['YTD Forms'].sum()), int(df_uni['YTD Admissions'].sum()), pct(df_uni['YTD Admissions'].sum(), df_uni['YTD Forms'].sum()),
        int(df_uni['MTD Forms'].sum()), int(df_uni['MTD Admissions'].sum()), pct(df_uni['MTD Admissions'].sum(), df_uni['MTD Forms'].sum()),
        int(df_uni['FTD Forms'].sum()), int(df_uni['FTD Admissions'].sum()), pct(df_uni['FTD Admissions'].sum(), df_uni['FTD Forms'].sum())
    ])
    style_worksheet(ws5)

    ws6 = wb.create_sheet("Unmatched_Admissions_Audit")
    ws6.append(['Status', 'Student ID', 'College', 'Supervisor', 'Counsellor', 'Amount', 'Admission Time IST'])
    if unmatched_adm_week.empty:
        ws6.append(['OK - no unmatched admissions', '', '', '', '', '', ''])
    else:
        for _, row in unmatched_adm_week.iterrows():
            ws6.append(['MISSING_IN_CONFIG', row.get('student_id'), row.get('college_name'), row.get('supervisor_name'), row.get('counsellor_name'), row.get('fee_deposit'), str(row.get('created_at'))])
    style_worksheet(ws6)

    filename = os.path.join(BASE_DIR, 'Daily_Online_LMS_Reports_V2.xlsx')
    prev_filename = os.path.join(BASE_DIR, 'Daily_Online_LMS_Reports_Prev.xlsx')
    
    if os.path.exists(filename):
        import shutil
        shutil.copy2(filename, prev_filename)
        
    wb.save(filename)
    print(f"Successfully generated {filename}")

if __name__ == "__main__":
    generate_reports()
